#!/usr/bin/env python3
"""
analyse_resultats_votation.py — Skill analyse-resultats-votation
Analyse les résultats d'une votation fédérale suisse par canton et par district
à partir du fichier resultats_par_domaine.xlsx (BFS).

Usage:
  python -u analyse_resultats_votation.py --xlsx sources/swissvotes/votation_86/resultats_par_domaine.xlsx --votation-id 86
  python -u analyse_resultats_votation.py --xlsx sources/swissvotes/votation_86/resultats_par_domaine.xlsx --votation-id 86 --json
"""

import argparse
import json
import os
import sys
from datetime import datetime
from typing import Optional

try:
    import openpyxl
except ImportError:
    print("[ERREUR] La bibliothèque openpyxl est requise. Installez-la avec : pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ---------------------------------------------------------------------------
# Chemins par défaut
# ---------------------------------------------------------------------------
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.abspath(os.path.join(_SCRIPT_DIR, "..", "..", "..", ".."))

# ---------------------------------------------------------------------------
# Correspondance numéro canton → nom français
# ---------------------------------------------------------------------------
CANTON_NAMES_FR = {
    1:  "Zurich",
    2:  "Berne",
    3:  "Lucerne",
    4:  "Uri",
    5:  "Schwytz",
    6:  "Obwald",
    7:  "Nidwald",
    8:  "Glaris",
    9:  "Zoug",
    10: "Fribourg",
    11: "Soleure",
    12: "Bâle-Ville",
    13: "Bâle-Campagne",
    14: "Schaffhouse",
    15: "Appenzell Rh.-Ext.",
    16: "Appenzell Rh.-Int.",
    17: "Saint-Gall",
    18: "Grisons",
    19: "Argovie",
    20: "Thurgovie",
    21: "Tessin",
    22: "Vaud",
    23: "Valais",
    24: "Neuchâtel",
    25: "Genève",
    26: "Jura",
}

# Demi-cantons (comptent pour 0.5 voix cantonale) — numéros officiels actuels
DEMI_CANTONS = {6, 7, 15, 16, 12, 13}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_float(val) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _safe_int(val) -> Optional[int]:
    f = _safe_float(val)
    return int(round(f)) if f is not None else None


def _fmt_pct(val: Optional[float]) -> str:
    if val is None:
        return "–"
    return f"{val:.2f}\u202f%"


def _fmt_num(val: Optional[int]) -> str:
    if val is None:
        return "–"
    # Séparateur de milliers avec espace insécable
    return f"{val:,}".replace(",", "\u202f")


def _vote_label(oui_pct: Optional[float]) -> str:
    if oui_pct is None:
        return "?"
    return "OUI ✓" if oui_pct > 50.0 else "NON ✗"


# ---------------------------------------------------------------------------
# Lecture de la feuille Kantone
# ---------------------------------------------------------------------------

def _parse_kantone(ws) -> dict:
    """
    Retourne un dict avec :
    - 'votation_num': int ou None
    - 'date_scrutin': str
    - 'titre': str
    - 'total': dict de chiffres nationaux
    - 'cantons': liste de dicts par canton
    """
    rows = list(ws.iter_rows(values_only=True))

    # Métadonnées (lignes 0 et 1, index Python)
    row0 = rows[0] if len(rows) > 0 else []
    row1 = rows[1] if len(rows) > 1 else []

    votation_num = _safe_int(row0[0]) if row0 else None
    date_scrutin = str(row0[1]).strip() if row0 and row0[1] else ""
    titre = str(row1[1]).strip() if row1 and row1[1] else ""

    # Chercher le total et les cantons
    total = None
    cantons = []

    for row in rows:
        if not any(v is not None for v in row):
            continue
        col_a = row[0]
        col_b = str(row[1]).strip() if row[1] is not None else ""

        # Ligne du total suisse
        if col_a is None and col_b.lower() in ("total", "total schweiz", "total suisse"):
            total = _extract_domain_row(None, col_b, row)
            continue

        # Ligne de canton (colonne A = entier 1..26)
        if isinstance(col_a, (int, float)) and col_a not in (0, 86) and col_b:
            num = int(col_a)
            if 1 <= num <= 26:
                cantons.append(_extract_domain_row(num, col_b, row))

    return {
        "votation_num": votation_num,
        "date_scrutin": date_scrutin,
        "titre": titre,
        "total": total,
        "cantons": cantons,
    }


def _extract_domain_row(num, nom, row) -> dict:
    """Extrait les colonnes C..K d'une ligne de données (index 2..10)."""
    inscrits      = _safe_int(row[2])
    bulletins     = _safe_int(row[3])
    participation = _safe_float(row[4])
    blancs        = _safe_int(row[5])
    invalides     = _safe_int(row[6])
    valables      = _safe_int(row[7])
    oui           = _safe_int(row[8])
    non           = _safe_int(row[9])
    oui_pct       = _safe_float(row[10])

    return {
        "numero":          num,
        "nom":             nom,
        "inscrits":        inscrits,
        "bulletins":       bulletins,
        "participation":   participation,
        "blancs":          blancs,
        "invalides":       invalides,
        "valables":        valables,
        "oui":             oui,
        "non":             non,
        "oui_pct":         oui_pct,
    }


# ---------------------------------------------------------------------------
# Lecture de la feuille Bezirke
# ---------------------------------------------------------------------------

def _parse_bezirke(ws) -> list:
    """
    Retourne une liste de dicts par district.
    Reconstitue l'appartenance cantonale par ordre d'apparition des cantons.
    """
    rows = list(ws.iter_rows(values_only=True))

    districts = []
    current_canton_num = None   # numéro de canton courant (reconstitué)
    current_canton_pos = -1     # position dans la liste des cantons de Kantone

    # Numéros de cantons dans l'ordre d'apparition attendu dans Bezirke
    # (même ordre que la feuille Kantone)
    canton_sequence = list(range(1, 27))
    canton_seq_idx = 0

    # La liste des districts par canton est contigüe : on detect le passage
    # au canton suivant dès qu'on voit une ligne dont la col A correspond
    # à un nom de canton (ligne de sous-titre, souvent avec seulement col A non nulle).
    # Dans le format actuel, les districts sont listés sans ligne de sous-titre par canton.
    # On reconstitue via les totaux par canton insérés ponctuellement.

    # Repérer les lignes contenant le total d'un canton : col A = nom du canton
    # et col B = nombre d'inscrits identique au total du canton.
    # → Simplement, on collecte toutes les lignes dont col A est une string non vide
    #   et qui ressemble à un nom de district (pas une ligne de titre/note).

    header_found = False
    for row in rows:
        col_a = row[0]

        # Chercher la ligne d'entête
        if not header_found:
            if col_a == "Bezirk":
                header_found = True
            continue

        if not isinstance(col_a, str) or not col_a.strip():
            continue

        nom = col_a.strip()

        # Ignorer les lignes de note (long texte)
        if len(nom) > 80:
            continue

        # Ignorer "Total Schweiz" ou équivalents
        if nom.lower() in ("total schweiz", "total suisse", "total"):
            continue

        inscrits  = _safe_int(row[1])
        bulletins = _safe_int(row[2])
        particip  = _safe_float(row[3])
        blancs    = _safe_int(row[4])
        invalides = _safe_int(row[5])
        valables  = _safe_int(row[6])
        oui       = _safe_int(row[7])
        non       = _safe_int(row[8])
        oui_pct   = _safe_float(row[9])

        districts.append({
            "nom":          nom,
            "inscrits":     inscrits,
            "bulletins":    bulletins,
            "participation": particip,
            "blancs":       blancs,
            "invalides":    invalides,
            "valables":     valables,
            "oui":          oui,
            "non":          non,
            "oui_pct":      oui_pct,
            # canton_num sera injecté lors de la reconstruction
            "canton_num":   None,
        })

    return districts


def _assign_canton_to_districts(districts: list, cantons: list) -> None:
    """
    Reconstitue l'appartenance des districts à leur canton.

    Stratégie : dans la feuille Bezirke, les districts sont regroupés par canton
    dans le même ordre que la feuille Kantone. On identifie le passage d'un canton
    au suivant en cumulant les électeurs inscrits des districts et en comparant
    ce cumul au total d'inscrits du canton courant (source BFS, feuille Kantone).

    Seuil de transition : 97 % du total cantonal (la source Linder couvre en général
    98–100 % des inscrits BFS, avec une marge de 1–3 % pour les différences de
    périmètre entre les deux sources).
    """
    if not cantons:
        return

    # Cumulation des inscrits par district ; on avance au canton suivant
    # dès que la somme atteint ≥ 97 % du total BFS du canton courant.
    canton_idx: int = 0
    cumul_inscrits: int = 0
    for d in districts:
        if canton_idx >= len(cantons):
            d["canton_num"] = cantons[-1]["numero"] if cantons else None
            continue
        canton = cantons[canton_idx]
        canton_total_inscrits = canton.get("inscrits") or 0
        d["canton_num"] = canton["numero"]
        if d.get("inscrits"):
            cumul_inscrits += d["inscrits"]
        if canton_total_inscrits > 0 and cumul_inscrits >= canton_total_inscrits * 0.97:
            canton_idx += 1
            cumul_inscrits = 0


# ---------------------------------------------------------------------------
# Construction du Markdown
# ---------------------------------------------------------------------------

def _render_markdown(
    votation_id: str,
    titre: str,
    date_scrutin: str,
    total: Optional[dict],
    cantons: list,
    districts: list,
    xlsx_path: str,
) -> str:
    today = datetime.now().strftime("%Y-%m-%d")
    titre_md = f"Résultats de la votation n°\u202f{votation_id}" + (f" — {titre}" if titre else "")

    lines = []

    # Front matter
    lines.append("---")
    lines.append(f'title: "{titre_md}"')
    lines.append(f'date_creation: "{today}"')
    lines.append('author: "skill analyse-resultats-votation"')
    lines.append(f'sources:')
    lines.append(f'  - "{xlsx_path}"')
    lines.append("---")
    lines.append("")
    lines.append(f"# {titre_md}")
    lines.append("")
    if date_scrutin:
        lines.append(f"**Date du scrutin** : {date_scrutin}")
        lines.append("")

    # --- Résultats nationaux ---
    lines.append("## Résultats nationaux")
    lines.append("")
    if total:
        oui_pct = total.get("oui_pct")
        resultat = "**accepté** ✓" if (oui_pct is not None and oui_pct > 50.0) else "**rejeté** ✗"
        lines.append(f"| Indicateur | Valeur |")
        lines.append(f"| ---------- | ------ |")
        lines.append(f"| Électeurs inscrits | {_fmt_num(total.get('inscrits'))} |")
        lines.append(f"| Bulletins déposés | {_fmt_num(total.get('bulletins'))} |")
        lines.append(f"| Participation | {_fmt_pct(total.get('participation'))} |")
        lines.append(f"| Bulletins blancs | {_fmt_num(total.get('blancs'))} |")
        lines.append(f"| Bulletins invalides | {_fmt_num(total.get('invalides'))} |")
        lines.append(f"| Suffrages valables | {_fmt_num(total.get('valables'))} |")
        lines.append(f"| Voix OUI | {_fmt_num(total.get('oui'))} |")
        lines.append(f"| Voix NON | {_fmt_num(total.get('non'))} |")
        lines.append(f"| % OUI | {_fmt_pct(oui_pct)} |")
        lines.append(f"| Résultat | {resultat} |")
    else:
        lines.append("_Données nationales non disponibles dans le fichier source._")
    lines.append("")

    # --- Résultats par canton ---
    lines.append("## Résultats par canton")
    lines.append("")
    cantons_sorted = sorted(
        [c for c in cantons if c.get("oui_pct") is not None],
        key=lambda x: x["oui_pct"],
        reverse=True,
    )
    if cantons_sorted:
        lines.append("| Canton | Inscrits | Participation | Suffrages valables | Voix OUI | Voix NON | % OUI | Résultat |")
        lines.append("| ------ | -------- | ------------- | ------------------ | -------- | -------- | ----- | -------- |")
        for c in cantons_sorted:
            nom_fr = CANTON_NAMES_FR.get(c["numero"], c["nom"])
            lines.append(
                f"| {nom_fr} "
                f"| {_fmt_num(c.get('inscrits'))} "
                f"| {_fmt_pct(c.get('participation'))} "
                f"| {_fmt_num(c.get('valables'))} "
                f"| {_fmt_num(c.get('oui'))} "
                f"| {_fmt_num(c.get('non'))} "
                f"| {_fmt_pct(c.get('oui_pct'))} "
                f"| {_vote_label(c.get('oui_pct'))} |"
            )
    else:
        lines.append("_Aucune donnée cantonale disponible._")
    lines.append("")

    # --- Majorité cantonale ---
    lines.append("## Majorité cantonale")
    lines.append("")
    oui_cantons = []
    non_cantons = []
    score_oui = 0.0
    score_non = 0.0
    for c in cantons:
        num = c.get("numero")
        oui_pct = c.get("oui_pct")
        if oui_pct is None or num is None:
            continue
        poids = 0.5 if num in DEMI_CANTONS else 1.0
        nom_fr = CANTON_NAMES_FR.get(num, c["nom"])
        if oui_pct > 50.0:
            oui_cantons.append(nom_fr)
            score_oui += poids
        else:
            non_cantons.append(nom_fr)
            score_non += poids

    def _fmt_score(s: float) -> str:
        return str(int(s)) if s == int(s) else str(s)

    lines.append(f"_Note : pour les initiatives populaires et les référendums obligatoires, une double majorité (peuple ET cantons) est requise._")
    lines.append("")
    lines.append(f"| | Nombre de voix cantonales |")
    lines.append(f"| --- | --- |")
    lines.append(f"| Cantons OUI | **{_fmt_score(score_oui)}** / {_fmt_score(score_oui + score_non)} |")
    lines.append(f"| Cantons NON | {_fmt_score(score_non)} / {_fmt_score(score_oui + score_non)} |")
    majority_cantons = score_oui > score_non
    lines.append(f"| Majorité cantonale | {'**acquise** ✓' if majority_cantons else '**non acquise** ✗'} |")
    lines.append("")
    if oui_cantons:
        lines.append(f"**Cantons OUI** ({len(oui_cantons)}) : {', '.join(oui_cantons)}")
        lines.append("")
    if non_cantons:
        lines.append(f"**Cantons NON** ({len(non_cantons)}) : {', '.join(non_cantons)}")
        lines.append("")

    # --- Résultats par district ---
    lines.append("## Résultats par district")
    lines.append("")
    real_districts = [
        d for d in districts
        if not d.get("_is_canton_total") and d.get("oui_pct") is not None
    ]
    districts_sorted = sorted(real_districts, key=lambda x: x["oui_pct"], reverse=True)
    if districts_sorted:
        lines.append("| District | Canton | Inscrits | Suffrages valables | Voix OUI | Voix NON | % OUI | Résultat |")
        lines.append("| -------- | ------ | -------- | ------------------ | -------- | -------- | ----- | -------- |")
        for d in districts_sorted:
            canton_num = d.get("canton_num")
            canton_fr = CANTON_NAMES_FR.get(canton_num, str(canton_num) if canton_num else "?")
            lines.append(
                f"| {d['nom']} "
                f"| {canton_fr} "
                f"| {_fmt_num(d.get('inscrits'))} "
                f"| {_fmt_num(d.get('valables'))} "
                f"| {_fmt_num(d.get('oui'))} "
                f"| {_fmt_num(d.get('non'))} "
                f"| {_fmt_pct(d.get('oui_pct'))} "
                f"| {_vote_label(d.get('oui_pct'))} |"
            )
    else:
        lines.append("_Aucune donnée de district disponible._")
    lines.append("")

    # --- Analyse géographique ---
    lines.append("## Analyse géographique")
    lines.append("")
    if cantons_sorted:
        top3 = cantons_sorted[:3]
        bot3 = cantons_sorted[-3:]
        lines.append(
            "**Cantons les plus favorables** : "
            + ", ".join(f"{CANTON_NAMES_FR.get(c['numero'], c['nom'])} ({_fmt_pct(c['oui_pct'])})" for c in top3)
        )
        lines.append("")
        lines.append(
            "**Cantons les plus opposés** : "
            + ", ".join(f"{CANTON_NAMES_FR.get(c['numero'], c['nom'])} ({_fmt_pct(c['oui_pct'])})" for c in reversed(bot3))
        )
        lines.append("")

    # Clivage linguistique Röstigraben
    romands = {22, 23, 24, 25, 10}  # Vaud, Valais, Neuchâtel, Genève, Fribourg (partiellement)
    tessinois = {21}
    germanophones = set(range(1, 26)) - romands - tessinois

    def _mean_oui(nums: set) -> Optional[float]:
        vals = [c["oui_pct"] for c in cantons if c.get("numero") in nums and c.get("oui_pct") is not None]
        return sum(vals) / len(vals) if vals else None

    moy_roman = _mean_oui(romands)
    moy_german = _mean_oui(germanophones)
    moy_tessin = _mean_oui(tessinois)

    if moy_roman is not None and moy_german is not None:
        diff = abs(moy_roman - moy_german)
        lines.append(
            f"**Clivage linguistique (Röstigraben)** : "
            f"cantons romands : {moy_roman:.1f}\u202f% OUI en moyenne ; "
            f"cantons alémaniques : {moy_german:.1f}\u202f% OUI en moyenne "
            f"(écart : {diff:.1f} pts)."
        )
        if diff >= 10:
            lines.append("→ Écart linguistique significatif (≥ 10 points) dans ce scrutin.")
        else:
            lines.append("→ Pas de Röstigraben marqué dans ce scrutin.")
        lines.append("")

    if moy_tessin is not None:
        lines.append(f"**Tessin** : {moy_tessin:.1f}\u202f% OUI en moyenne.")
        lines.append("")

    # District le plus favorable et le plus opposé
    if districts_sorted:
        top_d = districts_sorted[0]
        bot_d = districts_sorted[-1]
        lines.append(
            f"**District le plus favorable** : {top_d['nom']} "
            f"({CANTON_NAMES_FR.get(top_d.get('canton_num'), '?')}) "
            f"avec {_fmt_pct(top_d['oui_pct'])} OUI."
        )
        lines.append(
            f"**District le plus opposé** : {bot_d['nom']} "
            f"({CANTON_NAMES_FR.get(bot_d.get('canton_num'), '?')}) "
            f"avec {_fmt_pct(bot_d['oui_pct'])} OUI."
        )
        lines.append("")

    # Districts qui s'écartent le plus de la tendance de leur canton
    lines.append("### Districts s'écartant le plus de la tendance cantonale")
    lines.append("")
    canton_mean = {}
    for c in cantons:
        if c.get("numero") and c.get("oui_pct") is not None:
            canton_mean[c["numero"]] = c["oui_pct"]

    deviations = []
    for d in real_districts:
        num = d.get("canton_num")
        if num and num in canton_mean and d.get("oui_pct") is not None:
            dev = d["oui_pct"] - canton_mean[num]
            deviations.append((abs(dev), dev, d, num))

    deviations.sort(reverse=True, key=lambda x: x[0])
    if deviations:
        lines.append("| District | Canton | % OUI district | % OUI canton | Écart |")
        lines.append("| -------- | ------ | -------------- | ------------ | ----- |")
        for _, dev, d, num in deviations[:10]:
            signe = "+" if dev >= 0 else ""
            lines.append(
                f"| {d['nom']} "
                f"| {CANTON_NAMES_FR.get(num, str(num))} "
                f"| {_fmt_pct(d['oui_pct'])} "
                f"| {_fmt_pct(canton_mean[num])} "
                f"| {signe}{dev:.1f}\u202fpts |"
            )
    else:
        lines.append("_Impossible de calculer les écarts (appartenance cantonale non reconstituée)._")
    lines.append("")

    # --- Incertitudes ---
    lines.append("## Incertitudes et données manquantes")
    lines.append("")
    incertitudes = []
    if total is None:
        incertitudes.append("- Ligne de total national introuvable dans la feuille Kantone.")
    missing_canton_nums = [
        CANTON_NAMES_FR.get(i, str(i))
        for i in range(1, 27)
        if not any(c.get("numero") == i for c in cantons)
    ]
    if missing_canton_nums:
        incertitudes.append(f"- Cantons absents du fichier source : {', '.join(missing_canton_nums)}.")
    dist_no_canton = [d for d in real_districts if d.get("canton_num") is None]
    if dist_no_canton:
        incertitudes.append(
            f"- {len(dist_no_canton)} district(s) sans appartenance cantonale reconstituée : "
            + ", ".join(d["nom"] for d in dist_no_canton[:5])
            + ("…" if len(dist_no_canton) > 5 else "") + "."
        )
    dist_no_particip = [d for d in real_districts if d.get("bulletins") is None and d.get("inscrits") is not None]
    if dist_no_particip:
        incertitudes.append(
            f"- {len(dist_no_particip)} district(s) sans données de bulletins déposés (source Linder et al. 2007) : "
            "participation non calculable au niveau des districts."
        )
    if not incertitudes:
        incertitudes.append("- Aucune donnée manquante majeure détectée.")
    lines.extend(incertitudes)
    lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Point d'entrée
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Analyse les résultats d'une votation fédérale suisse par canton et district."
    )
    parser.add_argument("--xlsx", required=True, help="Chemin vers resultats_par_domaine.xlsx")
    parser.add_argument("--votation-id", default=None, help="Numéro de la votation (ex. 86)")
    parser.add_argument("--output", default=None, help="Chemin du fichier de sortie Markdown")
    parser.add_argument("--json", action="store_true", help="Produit aussi un fichier JSON")
    args = parser.parse_args()

    xlsx_path = args.xlsx
    if not os.path.isabs(xlsx_path):
        xlsx_path = os.path.join(_PROJECT_ROOT, xlsx_path)
    if not os.path.isfile(xlsx_path):
        print(f"[ERREUR] Fichier introuvable : {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    print(f"[INFO] Lecture de : {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    sheet_names_lower = {s.lower(): s for s in wb.sheetnames}

    # Feuille Kantone
    kantone_name = sheet_names_lower.get("kantone")
    if not kantone_name:
        print("[ERREUR] Feuille 'Kantone' introuvable dans le fichier xlsx.", file=sys.stderr)
        sys.exit(1)
    kantone_data = _parse_kantone(wb[kantone_name])

    # Feuille Bezirke
    bezirke_name = sheet_names_lower.get("bezirke")
    districts = []
    if bezirke_name:
        districts = _parse_bezirke(wb[bezirke_name])
        _assign_canton_to_districts(districts, kantone_data["cantons"])
    else:
        print("[AVERTISSEMENT] Feuille 'Bezirke' introuvable — données de districts ignorées.", file=sys.stderr)

    wb.close()

    # Numéro de votation
    votation_id = args.votation_id
    if not votation_id:
        votation_id = str(kantone_data.get("votation_num") or "inconnu")

    # Chemin de sortie
    if args.output:
        out_path = args.output
        if not os.path.isabs(out_path):
            out_path = os.path.join(_PROJECT_ROOT, out_path)
    else:
        out_dir = os.path.join(_PROJECT_ROOT, "sources", "swissvotes", f"votation_{votation_id}")
        os.makedirs(out_dir, exist_ok=True)
        out_path = os.path.join(out_dir, "resultats_par_domaine.md")

    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    # Rendu Markdown
    md = _render_markdown(
        votation_id=votation_id,
        titre=kantone_data.get("titre", ""),
        date_scrutin=kantone_data.get("date_scrutin", ""),
        total=kantone_data.get("total"),
        cantons=kantone_data.get("cantons", []),
        districts=districts,
        xlsx_path=xlsx_path,
    )

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(md)
    print(f"[OK] Fichier Markdown écrit : {out_path}")

    # JSON optionnel
    if args.json:
        out_json = os.path.splitext(out_path)[0] + ".json"
        total = kantone_data.get("total") or {}
        oui_pct_total = total.get("oui_pct")
        payload = {
            "generated": datetime.now().strftime("%Y-%m-%d"),
            "author": "skill analyse-resultats-votation",
            "votation_id": votation_id,
            "titre": kantone_data.get("titre", ""),
            "date_scrutin": kantone_data.get("date_scrutin", ""),
            "national": {
                "inscrits": total.get("inscrits"),
                "bulletins": total.get("bulletins"),
                "participation_pct": total.get("participation"),
                "blancs": total.get("blancs"),
                "invalides": total.get("invalides"),
                "valables": total.get("valables"),
                "oui": total.get("oui"),
                "non": total.get("non"),
                "oui_pct": oui_pct_total,
                "resultat": "accepté" if (oui_pct_total is not None and oui_pct_total > 50.0) else "rejeté",
            },
            "cantons": [
                {
                    "numero": c.get("numero"),
                    "nom_fr": CANTON_NAMES_FR.get(c.get("numero"), c.get("nom", "")),
                    "nom_source": c.get("nom"),
                    "inscrits": c.get("inscrits"),
                    "bulletins": c.get("bulletins"),
                    "participation_pct": c.get("participation"),
                    "blancs": c.get("blancs"),
                    "invalides": c.get("invalides"),
                    "valables": c.get("valables"),
                    "oui": c.get("oui"),
                    "non": c.get("non"),
                    "oui_pct": c.get("oui_pct"),
                    "vote": "OUI" if (c.get("oui_pct") is not None and c["oui_pct"] > 50.0) else "NON",
                }
                for c in sorted(kantone_data.get("cantons", []),
                                 key=lambda x: (x.get("oui_pct") or 0), reverse=True)
            ],
            "districts": [
                {
                    "nom": d.get("nom"),
                    "canton_num": d.get("canton_num"),
                    "canton_fr": CANTON_NAMES_FR.get(d.get("canton_num"), ""),
                    "inscrits": d.get("inscrits"),
                    "bulletins": d.get("bulletins"),
                    "participation_pct": d.get("participation"),
                    "valables": d.get("valables"),
                    "oui": d.get("oui"),
                    "non": d.get("non"),
                    "oui_pct": d.get("oui_pct"),
                    "vote": "OUI" if (d.get("oui_pct") is not None and d["oui_pct"] > 50.0) else "NON",
                }
                for d in sorted(
                    [x for x in districts if not x.get("_is_canton_total")],
                    key=lambda x: (x.get("oui_pct") or 0), reverse=True
                )
            ],
        }
        with open(out_json, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        print(f"[OK] Fichier JSON écrit : {out_json}")


if __name__ == "__main__":
    main()



